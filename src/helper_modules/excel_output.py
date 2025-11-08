from src.helper_modules.logger_setup import get_logger, shared_logger
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.layout import Layout, ManualLayout
from pathlib import Path
import pandas as pd 
from datetime import date

logger = get_logger(__name__)

def convert_to_excel(name: str, file: Path, *fs_dfs: pd.DataFrame, **ratio_dfs: pd.DataFrame) -> Path | None:
    """
    takes the file path provided by the user and creates an excel document with all the financial data of a listed company
    
    """

    file_name = f"output_{date.today()}_{name}.xlsx"
    file_path = file / file_name

    # make a list of sheet names and zip to match sheet name to financial statement
    sheet_names = ["balance_sheet", "income_statement", "cash_flows"]

    try:
        with pd.ExcelWriter(file_path) as writer:
            # iterate through all the provided dfs and use their name as the sheet name
            for df, sheet_name in zip(fs_dfs, sheet_names):
                df.to_excel(writer, sheet_name=sheet_name, header=False)
                logger.info(f"{sheet_name} successfully added to excel sheet")
                shared_logger.info(f"{sheet_name} successfully added to excel sheet")

            for key, value in ratio_dfs.items():
                value.sort_values("fiscalYear").to_excel(writer, sheet_name=key, index=False)
                logger.info(f"{key} successfully added to excel sheet")
                shared_logger.info(f"{key} successfully added to excel sheet")
                
            print(f"Excel output successfully created: {file_name}")
            logger.info(f"Excel output successfully created: {file_name}")
            shared_logger.info(f"Excel output successfully created: {file_name}")

            return file_path
    except Exception as e:
        print(e)
        logger.warning(f"{e}")
        shared_logger.warning(f"{e}")

def write_charts(workbook: Path) -> None:
    """
    takes an excel workbook and then adds charts on a new sheet for the financial ratios
    
    """

    # load the workbook and then create a sheet for the charts
    wb = load_workbook(workbook)
    logger.info(f"Workbook successfully loaded: {workbook}")
    shared_logger.info(f"Workbook successfully loaded: {workbook}")
  
    # iterate through the workbook and add a chart for each ratio sheet
    # parse all the ratio sheets from the workbook
    try:
        ratio_sheets = [ws for ws in wb.worksheets if "ratios" in ws.title.lower()]
        logger.info(f"Ratio sheets successfully parsed: {ratio_sheets}")
        shared_logger.info(f"Ratio sheets successfully parsed: {ratio_sheets}")
    except ValueError:
        print("No ratio charts were found in the workbook")
        logger.warning("No ratio charts were found in the workbook")
        shared_logger.warning("No ratio charts were found in the workbook")
        return None

    # create a chart sheet and overwrite if exists
    if "chart_sheet" in wb.sheetnames:
        wb.remove(wb["chart_sheet"])
        chart_ws = wb.create_sheet("chart_sheet")
    else:
        chart_ws = wb.create_sheet("chart_sheet")

    logger.info("Chart sheet has been added to the workbook")
    shared_logger.info("Chart sheet has been added to the workbook")

    # add charts to the chart sheet
    chart_row, chart_col = 1, 1
    charts_per_row = 2
    col_spacing = 10
    row_spacing = 15
    
    for i, ws in enumerate(ratio_sheets, start=1):
        if ws.max_row <= 2:
            print("Data provided is only for 1 period, therefore no charts created")
            logger.info("Data provided is only for 1 period, therefore no charts created")
            shared_logger.info("Data provided is only for 1 period, therefore no charts created")
            return None
        else:
            chart = LineChart()

            max_row = ws.max_row
            max_col = ws.max_column

            data = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=max_row)

            chart.title = ws.title
            chart.x_axis.title = "Year"
            chart.y_axis.title = "Ratios"

            anchor_cell = f"{get_column_letter(chart_col)}{chart_row}"

            chart.add_data(data, titles_from_data=True)

            # adjust the layout so that the axis labels and key do not overlap with data
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    x = .15, 
                    y = .25,
                    w = .75,
                    h = .65
                )
            )
            chart_ws.add_chart(chart, anchor_cell)

            # calculate the cell position for the next chart 
            if i % charts_per_row == 0:
                chart_col = 1
                chart_row += row_spacing
            else:
                chart_col += col_spacing

    wb.save(workbook)

    return None

    


